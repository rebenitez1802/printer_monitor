from openpyxl import Workbook

from openpyxl.styles import Border, Side, PatternFill, Font, GradientFill, Alignment,Side
from django.conf import settings
from openpyxl.drawing.image import Image

def style_range(ws, cell_range, border=Border(), fill=None, font=None, alignment=None):
    """
    Apply styles to a range of cells as if they were a single cell.

    :param ws:  Excel worksheet instance
    :param range: An excel range to style (e.g. A1:F20)
    :param border: An openpyxl Border
    :param fill: An openpyxl PatternFill or GradientFill
    :param font: An openpyxl Font object
    """

    top = Border(top=border.top)
    left = Border(left=border.left)
    right = Border(right=border.right)
    bottom = Border(bottom=border.bottom)

    first_cell = ws[cell_range.split(":")[0]]
    if alignment:
        ws.merge_cells(cell_range)
        first_cell.alignment = alignment

    rows = ws[cell_range]
    if font:
        first_cell.font = font

    for cell in rows[0]:
        cell.border = cell.border + top
    for cell in rows[-1]:
        cell.border = cell.border + bottom

    for row in rows:
        l = row[0]
        r = row[-1]
        l.border = l.border + left
        r.border = r.border + right
        if fill:
            for c in row:
                c.fill = fill


def generatePrinterReportXls(dataList,title):
	img = Image(settings.PROJ_PATH + '/printers/static/images/ixon-logo.png')

	filename = 'file.xlsx'
  	path = settings.PROJ_PATH + "/printers/static/download/" + filename
  	#print 'generating %s' % path
	wb = Workbook()
	ws = wb.active

	ws.column_dimensions['A'].width = 20
	ws.column_dimensions['B'].width = 19
	ws.column_dimensions['C'].width = 18
	ws.column_dimensions['D'].width = 18
	ws.column_dimensions['E'].width = 16
	ws.column_dimensions['F'].width = 26
	ws.row_dimensions[2].height = 80

	ws.merge_cells(start_row=2,start_column=1,end_row=2,end_column=4)
	
	ws.title = title
	ws['A2'] = 'Informe Mensual Recuento de Paginas'
	ws['A2'].font = Font(size=24)
	ws['A2'].alignment = Alignment(vertical='center')

	ws.add_image(img, 'E2')
	ws.append([])
	box = Border(left=Side(border_style='thick',
                           color='000000'),
                 right=Side(border_style='thick',
                            color='000000'),
                 top=Side(border_style='thick',
                          color='000000'),
                 bottom=Side(border_style='thick',
                             color='000000'))
	#ws.append(cols)
	offsetini = 3
	offsetfin = offsetini
	for data,backgroundStyles in dataList:
			
		
		for row in data:
			
			ws.append(row)
			offsetfin = offsetfin + 1

		if backgroundStyles:
			for s in backgroundStyles:
				index = s[0] + offsetini
				cells = ws['A%s'%index:'F%s'%index][0]
				for c in cells:
					
					c.fill =PatternFill("solid", fgColor=s[1])
		

		style_range(ws, 'A%s:F%s' % (offsetini+1,offsetfin), border=box)

		ws.append([''])
		
		offsetfin = offsetfin + 1
		offsetini = offsetfin
					#print c
		#print 'saving to disk file %s' % path
	try:
		
		wb.save(filename= path)
	except Exception as e:
		print(type(inst))    # the exception instance
		print(inst.args)     # arguments stored in .args
		print(inst)
	#print 'saved %s' % path
	return filename