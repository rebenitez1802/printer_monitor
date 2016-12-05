from openpyxl import load_workbook
from printers.models import Center, Printer
from django.core.exceptions import ObjectDoesNotExist

#from printers.scripts.migratePrinters import migratePrinters
#from printers.models import Customer 
#migratePrinters('/Users/Rafael/Downloads/Base de impresoras y centros MEGASALUD.xlsx',Customer.objects.get(pk=1))

def migratePrinters(xlsxPath, customer):
	wb = load_workbook(xlsxPath)
	if wb:
		i = 0
		for ws in wb.worksheets:
			i=i+1
			_center_name = ws.title
			try :
				center = Center.objects.get(name= _center_name, customer=customer.id)
			except ObjectDoesNotExist:
				center = Center(name=_center_name, customer_id=customer.id, order=i)
				center.save()
			if not center:
				center = Center(name=_center_name, customer_id=customer.id,order=i)
				center.save()
			if ws['B9'].value and 'MAC' in ws['B9'].value:

				print 'iterating %s' %_center_name
				
				for row in ws.iter_rows( min_row=10, max_row = ws.max_row - 1):
					if row[1].value:
						print 'printer: SN=%s, MAC=%s' % (row[2].value,row[1].value)
						'''
						center = models.ForeignKey(Center, on_delete=models.CASCADE)
						last_report = models.ForeignKey('PrinterReport', blank = True, null = True)
						brand = models.CharField(max_length=200)
						mac_address = models.CharField(max_length = 200, blank=True, null=True)
						model = models.CharField(max_length=200)
						serial_number = models.CharField(max_length=200, blank =True,null=True)
						host_name = models.CharField(max_length=200)
						ip_address = models.GenericIPAddressField(validators = [validate_ipv46_address])
						printer_type =  models.CharField(max_length=200, choices=PRINTER_TYPE, default = PRINTER_TYPE[0][0])
						location = models.CharField(max_length=200, blank=True, null=True)
						'''
						try:
							p = Printer.objects.get(center_id=center.id, serial_number =row[2].value)
						except ObjectDoesNotExist:
							try:
								p = Printer.objects.get(center_id=center.id, mac_address = row[1].value)
							except ObjectDoesNotExist:

								if '<No admitido>' in row[2].value:
									sn = None
								else:
									sn = row[2].value

								printer = Printer(
									center_id= center.id, 
									brand=row[5].value, 
									mac_address = row[1].value,
									model = row[6].value,
									serial_number= sn,
									host_name = row[3].value,
									ip_address = row[4].value,
									printer_type = 'mono',
									location=row[8].value )
								printer.save()
					
					